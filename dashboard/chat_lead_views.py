"""
dashboard/chat_lead_views.py
Chatbot Leads section for the custom dashboard.

This is a STANDALONE module — you do NOT need to edit dashboard/views.py.
Just drop this file into the dashboard/ app and wire the URLs (see urls snippet).
"""
import csv
import json

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone

from .views import staff_required   # reuse the existing decorator


# ── LIST ──────────────────────────────────────────────────────────────────────
@staff_required
def chat_lead_list(request):
    from chat.models import ChatLead
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')

    qs = ChatLead.objects.all()
    if status in ('pending', 'resolved'):
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(
            Q(name__icontains=search)  |
            Q(email__icontains=search) |
            Q(phone__icontains=search) |
            Q(query__icontains=search)
        )

    leads  = Paginator(qs.order_by('-created_at'), 15).get_page(request.GET.get('page'))
    counts = {
        'pending':  ChatLead.objects.filter(status='pending').count(),
        'resolved': ChatLead.objects.filter(status='resolved').count(),
        'all':      ChatLead.objects.count(),
    }
    return render(request, 'dashboard/chat_leads/list.html', {
        'leads': leads, 'status': status, 'search': search,
        'counts': counts, 'page_title': 'Chatbot Leads',
    })


# ── DETAIL (query + chat transcript + note) ───────────────────────────────────
@staff_required
def chat_lead_detail(request, pk):
    from chat.models import ChatLead, ChatSession
    obj = get_object_or_404(ChatLead, pk=pk)

    if request.method == 'POST':
        new_status = request.POST.get('status', obj.status)
        if new_status in ('pending', 'resolved'):
            obj.status = new_status
        obj.admin_note = request.POST.get('admin_note', '')
        obj.save()
        messages.success(request, 'Lead updated.')
        return redirect('dashboard:chat_lead_detail', pk=pk)

    chat_messages = []
    if obj.session_key:
        session = ChatSession.objects.filter(session_key=obj.session_key).first()
        if session:
            chat_messages = session.messages.order_by('created_at')

    return render(request, 'dashboard/chat_leads/detail.html', {
        'obj': obj, 'chat_messages': chat_messages, 'page_title': 'Lead Detail',
    })


# ── DELETE ────────────────────────────────────────────────────────────────────
@staff_required
def chat_lead_delete(request, pk):
    from chat.models import ChatLead
    obj = get_object_or_404(ChatLead, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Lead deleted.')
        return redirect('dashboard:chat_lead_list')
    return render(request, 'dashboard/confirm_delete.html', {'obj': obj, 'page_title': 'Delete Lead'})


# ── AJAX: inline status change (dropdown in list) ─────────────────────────────
@staff_required
@require_http_methods(['POST'])
def chat_lead_status(request, pk):
    from chat.models import ChatLead
    obj = get_object_or_404(ChatLead, pk=pk)
    try:
        data = json.loads(request.body.decode() or '{}')
    except json.JSONDecodeError:
        data = request.POST
    new_status = data.get('status', '')
    if new_status not in ('pending', 'resolved'):
        return JsonResponse({'ok': False, 'error': 'Invalid status'}, status=400)
    obj.status = new_status
    obj.save(update_fields=['status', 'updated_at'])
    return JsonResponse({'ok': True, 'status': new_status})


# ── EXPORT: Excel (openpyxl if installed, otherwise CSV) ──────────────────────
def _filtered_leads(request):
    from chat.models import ChatLead
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    qs = ChatLead.objects.all()
    if status in ('pending', 'resolved'):
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(
            Q(name__icontains=search)  | Q(email__icontains=search) |
            Q(phone__icontains=search) | Q(query__icontains=search)
        )
    return qs.order_by('-created_at')


EXPORT_HEADERS = ['Date', 'Name', 'Phone', 'Email', 'Query', 'Status', 'Session Key']


def _lead_row(l):
    return [
        timezone.localtime(l.created_at).strftime('%d-%m-%Y %H:%M'),
        l.name, l.phone, l.email, l.query,
        l.get_status_display(), l.session_key,
    ]


@staff_required
def chat_lead_export_excel(request):
    qs    = _filtered_leads(request)
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    try:
        # Real .xlsx if openpyxl is installed (pip install openpyxl)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = 'Chatbot Leads'
        ws.append(EXPORT_HEADERS)
        header_fill = PatternFill('solid', fgColor='F97316')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        for l in qs:
            ws.append(_lead_row(l))
        for col, width in zip('ABCDEFG', [18, 22, 15, 28, 50, 12, 30]):
            ws.column_dimensions[col].width = width
        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="chatbot_leads_{stamp}.xlsx"'
        wb.save(resp)
        return resp
    except ImportError:
        # Fallback: CSV (opens directly in Excel)
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="chatbot_leads_{stamp}.csv"'
        resp.write('\ufeff')  # BOM so Excel detects UTF-8
        writer = csv.writer(resp)
        writer.writerow(EXPORT_HEADERS)
        for l in qs:
            writer.writerow(_lead_row(l))
        return resp


# ── EXPORT: PDF (printable page — browser Save as PDF; zero dependencies) ─────
@staff_required
def chat_lead_export_pdf(request):
    qs = _filtered_leads(request)
    return render(request, 'dashboard/chat_leads/print.html', {
        'leads': qs,
        'generated_at': timezone.localtime(),
        'status': request.GET.get('status', ''),
        'search': request.GET.get('search', ''),
    })
